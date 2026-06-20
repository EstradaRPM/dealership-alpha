"""Generate the Tier / Macro-Loop planning workbook.

A formatted .xlsx scaffold for designing the dealership-sim progression spine.
Pre-seeds known facts from data/tier-progression.json + data/staff-roles.json,
and surfaces design axes the solo dev may not be considering.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join("docs", "planning", "tier-macro-loop-planner.xlsx")

# ---- palette ----
NAVY      = "1F2A44"   # sheet title bands
SECTION   = "2E5A88"   # category section rows
TIERHEAD  = "0F766E"   # tier column headers
SEED      = "E6F4EA"   # pre-filled known facts (light green)
FILLME    = "FFFFFF"   # cells for the user
PROMPT    = "FFF7E6"   # prompt / question cells (light amber)
ALT       = "F4F6F9"   # zebra
WHITE     = "FFFFFF"

thin = Side(style="thin", color="C9D2DD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style(ws, cell, *, bold=False, size=11, color="1B1B1B", fill=None,
          wrap=True, align="left", valign="top", italic=False):
    c = ws[cell]
    c.font = Font(bold=bold, size=size, color=color, italic=italic)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(wrap_text=wrap, horizontal=align, vertical=valign)
    c.border = border
    return c

def title_band(ws, text, span):
    ws.merge_cells(f"A1:{span}1")
    style(ws, "A1", bold=True, size=15, color="FFFFFF", fill=NAVY,
          align="left", valign="center")
    ws["A1"] = text
    ws.row_dimensions[1].height = 30

wb = Workbook()

# =====================================================================
# SHEET 1 — START HERE
# =====================================================================
ws = wb.active
ws.title = "Start Here"
ws.sheet_view.showGridLines = False
for col, w in {"A": 3, "B": 40, "C": 70, "D": 30}.items():
    ws.column_dimensions[col].width = w
title_band(ws, "Tier / Macro-Loop Planner  —  dealership sim", "D")

rows = [
    ("", ""),
    ("HOW TO USE THIS", "header"),
    ("1. Tier Map", "The main sheet. Each ROW is a design dimension; each COLUMN is a tier. "
        "Cells with a green tint are facts already locked in your code/spec — react to them, don't re-derive. "
        "White cells are yours to fill. Amber cells are open questions to resolve."),
    ("2. Macro Loop Canvas", "Answer the 5 prompts here FIRST. The whole tier structure falls out of "
        "prompt #2 (first hour vs. last hour of a long save). Do this before filling the Tier Map."),
    ("3. Staff & Unlocks", "Which role unlocks which verb/mechanic, and at which tier. Pre-seeded from staff-roles.json."),
    ("4. Open Questions", "Park every fork you hit so none get lost. The big one is already in there (tier ceiling / multi-store)."),
    ("5. Locked (ref)", "What's already DECIDED — the micro day-loop and the 4 UI forks from this session. Don't re-open these."),
    ("", ""),
    ("ORDER OF OPERATIONS", "header"),
    ("Step 1", "Fill the Macro Loop Canvas (5 prompts). Especially #2."),
    ("Step 2", "Decide the tier CEILING in Open Questions (how many tiers; is multi-store in the shippable product)."),
    ("Step 3", "Fill the Tier Map left-to-right, one tier column at a time. Don't try to fill all dimensions for all tiers at once."),
    ("Step 4", "Hand it back to Claude (see 'How to resume' below)."),
    ("", ""),
    ("AXES YOU MIGHT NOT BE CONSIDERING", "header"),
    ("(these are baked into the Tier Map rows on purpose — domain expertise covers the retail side; these are the GAME-DESIGN + scaling axes that are easy to skip)", "sub"),
    ("Floorplan financing & carrying cost", "Real dealers finance inventory on a floorplan line; carrying cost is a scaling pressure. Is the credit line itself a tier-gated lever?"),
    ("Independent → franchised transition", "Used-only independent lot vs. franchised new-car dealer (OEM allocation, holdback, incentives, objectives) is arguably the single biggest real progression axis. Is it a tier? A separate track?"),
    ("What gets AUTOMATED / removed each tier", "Your stated goal: graduate out of tedium. Explicitly map what stops needing manual attention at each tier — not just what's added."),
    ("New verbs vs. bigger numbers", "Each tier should add a new DECISION (a verb), not just larger inventory/cash. If a tier only scales numbers, it's a difficulty knob, not progression."),
    ("Dominant optimization per tier", "What is the player mainly trying to get right at this tier? If it's the same as the prior tier, the tiers may need merging."),
    ("Pacing / time-to-traverse", "Roughly how long (in-game and real) should a tier last? Guards against grind and against tiers flashing by."),
    ("Difficulty source per tier", "What makes THIS tier hard (cash flow? objectives? competition? regulation?) — should shift as you climb, not just intensify."),
    ("North star / ending", "Is there a win/retire/exit, or open-ended sandbox? Determines whether a ceiling even exists."),
    ("Failure & recovery per tier", "Already specced (T1 terminal -> T2 contraction -> T3+ recoverable). Confirm it still fits your redesigned loop."),
    ("Onboarding alignment", "Tier 1 must teach the match skill (demand-read + stock-to-match). Onboarding is a tier-1 design constraint, not a bolt-on."),
    ("", ""),
    ("HOW TO RESUME WITH CLAUDE", "header"),
    ("Easiest", "Open a new session in this repo and say: \"Read memory ui-mapping-pipeline and macro-loop-redesign-open, "
        "then let's resume the macro-loop grill — here's my filled planner.\" Then paste cells, attach the .xlsx, or send a photo."),
    ("What Claude will do", "Read the two memory files (they hold every locked decision + the paused grill state), ingest your filled planner, "
        "turn it into the progression spine, update the spec, THEN resume the Home-screen UI forks on top of it."),
    ("If you only have paper", "A photo of the paper works — Claude reads images. Or transcribe the 5 canvas answers + tier count into chat."),
    ("Where this file lives", "docs/planning/tier-macro-loop-planner.xlsx  (commit it so it travels with the repo and any session can find it)."),
]
r = 3
for label, val in rows:
    if val == "header":
        ws.merge_cells(f"B{r}:D{r}")
        style(ws, f"B{r}", bold=True, size=12, color="FFFFFF", fill=SECTION)
        ws[f"B{r}"] = label
        ws.row_dimensions[r].height = 22
    elif val == "sub":
        ws.merge_cells(f"B{r}:D{r}")
        style(ws, f"B{r}", italic=True, size=9, color="6B7280")
        ws[f"B{r}"] = label
    elif label == "" and val == "":
        ws.row_dimensions[r].height = 6
    else:
        style(ws, f"B{r}", bold=True, fill=ALT, valign="top")
        ws[f"B{r}"] = label
        ws.merge_cells(f"C{r}:D{r}")
        style(ws, f"C{r}")
        ws[f"C{r}"] = val
        ws.row_dimensions[r].height = max(30, 14 * (1 + len(val)//70))
    r += 1
ws.freeze_panes = "A2"

# =====================================================================
# SHEET 2 — TIER MAP  (the core)
# =====================================================================
ws = wb.create_sheet("Tier Map")
ws.sheet_view.showGridLines = False
N_TIERS = 6
last_col = get_column_letter(2 + N_TIERS)  # A=dim, B..= tiers, then notes
notes_col = get_column_letter(3 + N_TIERS)
title_band(ws, "Tier Map  —  rows = dimensions · columns = tiers · green = already locked · amber = open", notes_col)

# header row (row 2)
ws.column_dimensions["A"].width = 30
style(ws, "A2", bold=True, color="FFFFFF", fill=TIERHEAD)
ws["A2"] = "Dimension  (down)   /   Tier  (across)"
tier_seed_labels = {
    1: "Tier 1\nGravel Yard",
    2: "Tier 2\nPaved Lot",
    3: "Tier 3\nSmall Showroom",
    4: "Tier 4\nFranchise? (spec: post-launch)",
    5: "Tier 5\nMulti-store? (spec: post-launch)",
    6: "Tier 6\n(spare — rename or delete)",
}
for t in range(1, N_TIERS + 1):
    col = get_column_letter(1 + t)
    ws.column_dimensions[col].width = 24
    style(ws, f"{col}2", bold=True, color="FFFFFF", fill=TIERHEAD, align="center", valign="center")
    ws[f"{col}2"] = tier_seed_labels[t]
ws.column_dimensions[notes_col].width = 40
style(ws, f"{notes_col}2", bold=True, color="FFFFFF", fill=TIERHEAD)
ws[f"{notes_col}2"] = "Open questions / notes"
ws.row_dimensions[2].height = 34

# (dimension, {tier_index: seeded_value}, note)
SECTIONS = [
    ("IDENTITY & FANTASY", [
        ("Tier label / name", {1: "Gravel Yard", 2: "Paved Lot", 3: "Small Showroom",
                               4: "Franchise (spec)", 5: "Multi-store (spec)"}, ""),
        ("Player fantasy / feel", {}, "What does owning this tier FEEL like?"),
        ("Franchise status", {1: "Independent / used", 2: "Independent / used"},
            "When (if ever) do you become a franchised NEW-car dealer? Big axis."),
        ("Narrative beat (chapter card)", {}, "Chapter-card moment on reaching this tier."),
    ]),
    ("FACILITY", [
        ("Lot / building type", {1: "Gravel yard", 2: "Paved lot", 3: "Small showroom"}, ""),
        ("Display capacity (units shown)", {}, ""),
        ("Service bays", {1: "0", 2: "(service unlocks T2)"}, "Service dept unlocks at Tier 2."),
        ("Showroom / indoor space", {1: "None", 3: "Yes"}, ""),
    ]),
    ("CAPITAL & FINANCE", [
        ("Entry cash / starting capital", {}, "Starting cash for a fresh career = Tier 1 value."),
        ("Floorplan credit line (inventory financing)", {}, "Is the line size tier-gated? Carrying cost scales with it."),
        ("Rent / mortgage / overhead", {}, ""),
        ("Capex cost to ADVANCE to next tier", {}, "What does the upgrade itself cost (the Growth/Store spend)?"),
    ]),
    ("INVENTORY", [
        ("Max inventory (units)", {}, ""),
        ("New vs Used access", {1: "Used only"}, "Tied to franchise status above."),
        ("Brand / OEM access", {}, "Which brands can you stock/sell at this tier?"),
        ("Segment access (trucks/luxury/etc.)", {}, ""),
    ]),
    ("STAFF & DELEGATION", [
        ("Headcount cap", {}, ""),
        ("Roles hireable (new this tier)", {1: "lot-porter, salesperson",
                                            2: "f&i-manager, used-car-manager, service-advisor",
                                            3: "gm"},
            "From staff-roles.json hireTiers. NOTE: 'bdc-rep' role does NOT exist yet (needed for appointments)."),
        ("Org depth / delegation available", {}, "How many layers can you delegate through?"),
        ("Automation level (% routine auto-resolved)", {1: "Low (hands-on)"},
            "YOUR KEY AXIS: how much tedium is gone by this tier?"),
    ]),
    ("DEPARTMENTS ACTIVE", [
        ("Sales", {1: "Yes"}, ""),
        ("BDC / follow-up", {1: "Basic (free)"}, "Booked-appointments gate behind a bdc-rep hire."),
        ("Service", {1: "No", 2: "Yes"}, "Unlocks Tier 2."),
        ("Full F&I (product shelf)", {1: "VSC+GAP only", 2: "Full (w/ F&I mgr)"},
            "Full shelf unlocks with F&I Manager hire."),
        ("Parts / Bodyshop", {1: "No", 3: "Yes"}, "Bodyshop unlocks at Tier 3 (collision mirror of Service)."),
    ]),
    ("VERBS / MECHANICS UNLOCKED (new decisions)", [
        ("New player VERB added this tier", {}, "The new DECISION, not just bigger numbers. Critical row."),
        ("Pricing / strategy levers", {}, ""),
        ("Marketing channels (demand influence)", {}, ""),
        ("Trade-ins / wholesale / auction", {}, ""),
        ("Special / subprime finance", {}, ""),
    ]),
    ("WHAT GETS AUTOMATED / REMOVED", [
        ("Tedium removed this tier", {}, "Explicitly: what stops needing manual attention? (your stated goal)"),
        ("What the player STOPS doing", {}, ""),
    ]),
    ("MARKET & DEMAND", [
        ("Customer volume ceiling / day", {}, ""),
        ("Walk-in vs appointment mix", {}, "Appointments scale with BDC investment."),
        ("Geographic draw / market reach", {}, ""),
        ("Competition exposure", {}, "More/tougher competitors as you grow?"),
    ]),
    ("OEM & OBJECTIVES", [
        ("OEM objective scale (the goals system)", {}, "4-band period objectives + stair-step bonus (locked)."),
        ("Holdback / incentives / allocation", {}, "Only meaningful once franchised."),
    ]),
    ("RISK & FAILURE", [
        ("Regulatory scrutiny level", {}, ""),
        ("Failure mode", {1: "Terminal (game over)", 2: "Contraction (debt overhang)",
                          3: "Recoverable (consent decree)"}, "Already specced — confirm it fits."),
    ]),
    ("PROGRESSION", [
        ("Gate to REACH this tier", {2: "$125k cash · 100 served · rep 62",
                                     3: "$400k cash · 300 served · rep 75"},
            "From tier-progression.json (checked every 28 days)."),
        ("Dominant optimization at this tier", {}, "What is the player mainly trying to get right here?"),
        ("Est. time to traverse (in-game / real)", {}, "Pacing guard against grind."),
        ("Difficulty source (what makes it hard)", {}, "Should shift across tiers, not just intensify."),
    ]),
]

r = 3
zebra = False
for sec_name, dims in SECTIONS:
    ws.merge_cells(f"A{r}:{notes_col}{r}")
    style(ws, f"A{r}", bold=True, size=11, color="FFFFFF", fill=SECTION)
    ws[f"A{r}"] = sec_name
    ws.row_dimensions[r].height = 20
    r += 1
    for dim, seeds, note in dims:
        rowfill = ALT if zebra else WHITE
        zebra = not zebra
        style(ws, f"A{r}", bold=True, fill=rowfill)
        ws[f"A{r}"] = dim
        for t in range(1, N_TIERS + 1):
            col = get_column_letter(1 + t)
            val = seeds.get(t, "")
            cellfill = SEED if val else rowfill
            style(ws, f"{col}{r}", fill=cellfill, align="left")
            ws[f"{col}{r}"] = val
        style(ws, f"{notes_col}{r}", fill=PROMPT if note else rowfill, italic=bool(note), color="8A6D1A" if note else "1B1B1B")
        ws[f"{notes_col}{r}"] = note
        ws.row_dimensions[r].height = 30
        r += 1

ws.freeze_panes = "B3"

# =====================================================================
# SHEET 3 — STAFF & UNLOCKS
# =====================================================================
ws = wb.create_sheet("Staff & Unlocks")
ws.sheet_view.showGridLines = False
cols = ["Role", "Department", "Hireable at tier", "Skills granted",
        "Mechanic / verb it unlocks", "Tedium it removes", "Notes"]
widths = [20, 14, 14, 30, 30, 28, 30]
title_band(ws, "Staff & Unlocks  —  which hire opens which verb (seeded from staff-roles.json)", get_column_letter(len(cols)))
for i, (c, w) in enumerate(zip(cols, widths), start=1):
    L = get_column_letter(i)
    ws.column_dimensions[L].width = w
    style(ws, f"{L}2", bold=True, color="FFFFFF", fill=TIERHEAD)
    ws[f"{L}2"] = c
ws.row_dimensions[2].height = 22

staff = [
    ("lot-porter", "—", "1 (start)", "productivity", "(feeder role)", "", "Promotes to salesperson/technician"),
    ("salesperson", "sales", "1 (start)", "product_knowledge, communication, rapport", "Floor sales auto-resolve", "Manual greet/qualify", ""),
    ("bdc-rep  ⟵ NEW", "bdc", "(you decide)", "(define)", "Booked APPOINTMENTS (this session's fork)", "Manual follow-up triage", "Does NOT exist in data yet — add it."),
    ("f&i-manager", "sales", "2", "finance_structuring, product_presentation", "Full F&I product shelf + F&I profit", "—", "Auto-resolved F&I, gated on this hire"),
    ("used-car-manager", "sales", "2", "condition_reading, pricing", "(define — recon/appraisal authority?)", "Manual appraisal?", ""),
    ("service-advisor", "service", "2 (dept)", "diagnostic_clarity, write_up, upsell", "Service write-ups auto-resolve", "Manual service intake", ""),
    ("service-manager", "service", "(promotion)", "shop_throughput, warranty", "Service dept delegation", "Manual dispatch", ""),
    ("sales-manager", "sales", "(promotion)", "pricing, t_o_closing", "Desk/T.O. authority", "Manual deal desking", ""),
    ("gm", "—", "3", "(none — trait-gated)", "Advanced KPIs (PVR, PPRU); full delegation", "Most day-to-day oversight", "Top of the org DAG"),
]
r = 3
for i, row in enumerate(staff):
    rowfill = ALT if i % 2 else WHITE
    for j, val in enumerate(row, start=1):
        L = get_column_letter(j)
        fill = SEED if (j == 1 and "NEW" not in val) else rowfill
        if "NEW" in str(row[0]):
            fill = PROMPT
        style(ws, f"{L}{r}", fill=fill, bold=(j == 1))
        ws[f"{L}{r}"] = val
    ws.row_dimensions[r].height = 30
    r += 1
ws.freeze_panes = "A3"

# =====================================================================
# SHEET 4 — MACRO LOOP CANVAS
# =====================================================================
ws = wb.create_sheet("Macro Loop Canvas")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 80
title_band(ws, "Macro Loop Canvas  —  answer these FIRST (especially #2)", "C")
prompts = [
    ("1 · NORTH STAR", "What is the player ultimately chasing? Survive & retire rich? Biggest name in the metro? "
        "A narrative ending? Open-ended sandbox? (Decides whether a 'ceiling' even exists.)"),
    ("2 · TRANSFORMATION  ⟵ start here", "Describe the FIRST hour vs. the LAST hour of a 40-hour save. "
        "What is the player DOING in each? The progression unit usually falls right out of this answer."),
    ("3 · PROGRESSION UNIT", "What actually advances, and what does advancing CHANGE about play? "
        "Tiers were one answer. Could be: reputation/brand access, departments, capital/credit, automation depth, locations. "
        "What unlocks new VERBS (not just bigger numbers)?"),
    ("4 · THREAT / TENSION", "What keeps decisions tense so they matter? Cash-flow/bankruptcy? OEM objectives? "
        "Competition? Regulation? Pick the ONE or TWO that are the spine of the difficulty."),
    ("5 · SESSION SHAPE", "What's one satisfying sitting (10 min? a full in-game month?), and what's the hook "
        "that pulls the player back tomorrow?"),
    ("EARLY GAME — what the player does", ""),
    ("MID GAME — what changes", ""),
    ("LATE GAME — what the player does", ""),
]
r = 3
for label, hint in prompts:
    style(ws, f"B{r}", bold=True, size=12, color="FFFFFF", fill=SECTION, valign="center")
    ws[f"B{r}"] = label
    style(ws, f"C{r}", italic=True, color="6B7280")
    ws[f"C{r}"] = hint
    ws.row_dimensions[r].height = 18 if hint == "" else 42
    r += 1
    # answer space
    style(ws, f"B{r}", fill=ALT)
    ws[f"B{r}"] = "your answer →"
    style(ws, f"C{r}", fill=FILLME)
    ws[f"C{r}"] = ""
    ws.row_dimensions[r].height = 70
    r += 1
ws.freeze_panes = "A2"

# =====================================================================
# SHEET 5 — OPEN QUESTIONS
# =====================================================================
ws = wb.create_sheet("Open Questions")
ws.sheet_view.showGridLines = False
cols = ["#", "Question", "Options", "Leaning", "Decision", "Status"]
widths = [5, 38, 40, 22, 28, 14]
title_band(ws, "Open Questions  —  park every fork here", get_column_letter(len(cols)))
for i, (c, w) in enumerate(zip(cols, widths), start=1):
    L = get_column_letter(i)
    ws.column_dimensions[L].width = w
    style(ws, f"{L}2", bold=True, color="FFFFFF", fill=TIERHEAD)
    ws[f"{L}2"] = c
ws.row_dimensions[2].height = 22
qs = [
    ("Q1", "Tier CEILING: how many tiers, and is multi-store in the SHIPPABLE product?",
     "A) cap single-enterprise (T1-3/4), multi-store post-ship  ·  B) commit multi-store (T5) now",
     "A (solo-dev realism)", "", "OPEN"),
    ("Q2", "Independent→franchised: is becoming a NEW-car franchise a tier, a separate track, or out?",
     "tier axis / parallel track / cut", "", "", "OPEN"),
    ("Q3", "Is there an ENDING (retire/exit) or open-ended sandbox?",
     "ending / sandbox / both", "", "", "OPEN"),
    ("Q4", "Does each tier add a new VERB, or mostly scale numbers?",
     "verb-per-tier / scale-only", "verb-per-tier", "", "OPEN"),
    ("Q5", "bdc-rep role: define its hireTier + skills (needed for appointments).",
     "", "", "", "OPEN"),
]
r = 3
for i, row in enumerate(qs):
    rowfill = ALT if i % 2 else WHITE
    for j, val in enumerate(row, start=1):
        L = get_column_letter(j)
        style(ws, f"{L}{r}", fill=PROMPT if j in (2,) else rowfill, bold=(j == 1))
        ws[f"{L}{r}"] = val
    ws.row_dimensions[r].height = 40
    r += 1
ws.freeze_panes = "A3"

# =====================================================================
# SHEET 6 — LOCKED (ref)
# =====================================================================
ws = wb.create_sheet("Locked (ref)")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 90
title_band(ws, "Locked decisions (reference — do not re-open)", "C")
locked = [
    ("MICRO DAY-LOOP", "header"),
    ("One day", "Read demand 'weather' → stock & price inventory to match → press START DAY → "
        "watch the floor resolve in real time, respond only to exceptions → day recap. "
        "Managerial-watch, not click-every-sale. F&I auto-resolves."),
    ("", ""),
    ("UI FORKS LOCKED THIS SESSION", "header"),
    ("1 · Navigation", "Floor is a MODE (entered via START DAY), not a tab. Bottom tabs = Home · Operations · People · "
        "Finance · Growth. Departments are rooms launched from Home. Marketing→Growth, Sales→Operations, Store/capex→Growth. "
        "[Revisit if macro-loop changes the multi-store assumption.]"),
    ("2 · Goals/Targets", "Progression-tied period objectives (OEM/tier), daily figures = pace. 4-band gradient "
        "(Exceed/Meet/Near-miss/Miss) + persistence escalation reusing tier failure paths + stair-step retroactive bonus."),
    ("3 · Weather", "Real demand input. Season→mix (winter ↑trucks/SUV ↓sports; summer flips). Daily weather→volume "
        "(nice ↑, bad ↓). Small magnitudes in data/. Seeded/deterministic + forecast. Wires into DemandShaper + CustomerPool."),
    ("4 · Appointments", "Real booked-appointments as a FollowUpPool extension, gated behind a NEW bdc-rep role. "
        "BDC → booked → scheduled high-intent floor arrival + show/no-show. Basic follow-up stays free."),
    ("", ""),
    ("FLOOR VIEW FEASIBILITY", "header"),
    ("Stack", "Buildable in current RN+Expo as 2.5D: baked isometric backdrop (AI render as static asset) + live 2D token "
        "overlay driven by FloorSim. NO Unity. For polish use react-native-skia on a custom Expo dev build. "
        "Token positions are a VIEW concern (FloorSim emits state, not coordinates)."),
]
r = 3
for label, val in locked:
    if val == "header":
        ws.merge_cells(f"B{r}:C{r}")
        style(ws, f"B{r}", bold=True, size=12, color="FFFFFF", fill=SECTION)
        ws[f"B{r}"] = label
        ws.row_dimensions[r].height = 22
    elif label == "" and val == "":
        ws.row_dimensions[r].height = 6
    else:
        style(ws, f"B{r}", bold=True, fill=ALT)
        ws[f"B{r}"] = label
        style(ws, f"C{r}", fill=SEED)
        ws[f"C{r}"] = val
        ws.row_dimensions[r].height = max(34, 13 * (1 + len(val)//90))
    r += 1
ws.freeze_panes = "A2"

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("wrote", OUT, "with sheets:", wb.sheetnames)
